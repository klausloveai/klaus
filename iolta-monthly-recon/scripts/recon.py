#!/usr/bin/env python3
"""
IOLTA #3618 monthly bank reconciliation engine.

    python3 recon.py --stmt ~/Downloads/stmt.csv --month 2026-09 [--write] [--json out.json]

Reads a BoA statement (Business Advantage 360 CSV export, or `pdftotext -layout` output of an
eStatement), matches every line against Account-Journal.xlsx, and reports what tied and what did not.
With --write it backfills columns M (Cleared Date) / N (Statement Month). Without it, dry run.

NEVER plugs a difference. Anything that does not tie is reported, not adjusted.
"""
import argparse, csv, datetime, json, os, re, shutil, sys
from collections import defaultdict

try:
    import openpyxl
    from openpyxl.styles import Alignment
except ImportError:
    sys.exit("need openpyxl:  pip3 install openpyxl")

JOURNAL = os.path.expanduser(
    "~/Library/CloudStorage/GoogleDrive-klaus@lingtulaw.com/My Drive/Lingtu Law-Disbursement/"
    "IOLTA#3618/Account-Journal.xlsx")
SHEET = "Account journal"
HDR_ROW, FIRST_ROW = 9, 10
C = dict(date=1, payee=2, method=3, ck=4, purpose=5, dep=6, dis=7, bal=8,
         cli=9, notes=10, recL=11, recB=12, cleared=13, stmtmo=14)

num = lambda v: v if isinstance(v, (int, float)) else 0.0
c2  = lambda x: round(x + 0.0, 2)
def fmt(v):
    if isinstance(v, datetime.datetime): return v.strftime("%m/%d/%Y")
    return "" if v is None else str(v)

# ---------------------------------------------------------------- journal
def load_journal(path=JOURNAL):
    ws = openpyxl.load_workbook(path, data_only=True)[SHEET]
    last = FIRST_ROW - 1
    for r in range(FIRST_ROW, ws.max_row + 1):          # scan forward; rows 1-9 are merged, never touch
        if any(ws.cell(r, c).value is not None for c in range(1, 13)):
            last = r
    rows = []
    for r in range(FIRST_ROW, last + 1):
        ck = ws.cell(r, C["ck"]).value
        rows.append(dict(
            row=r, date=ws.cell(r, C["date"]).value, payee=ws.cell(r, C["payee"]).value,
            ck="" if ck is None else str(ck).strip(), purpose=ws.cell(r, C["purpose"]).value,
            dep=num(ws.cell(r, C["dep"]).value), dis=num(ws.cell(r, C["dis"]).value),
            cli=ws.cell(r, C["cli"]).value, notes=ws.cell(r, C["notes"]).value,
            cleared=ws.cell(r, C["cleared"]).value, stmtmo=ws.cell(r, C["stmtmo"]).value))
    return rows, last

def ensure_columns(path=JOURNAL):
    """Idempotent: add M/N headers if absent."""
    import copy
    wb = openpyxl.load_workbook(path, data_only=False); ws = wb[SHEET]
    if ws.cell(HDR_ROW, 13).value == "Cleared Date": return False
    src = ws.cell(HDR_ROW, 12)
    for col, title, w in ((13, "Cleared Date", 14.0), (14, "Statement Month", 15.5)):
        c = ws.cell(HDR_ROW, col); c.value = title
        for attr in ("font", "fill", "border", "alignment"):
            setattr(c, attr, copy.copy(getattr(src, attr)))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.auto_filter.ref = "A9:N15"
    wb.save(path); return True

# ---------------------------------------------------------------- statement
def parse_csv(path):
    out = []
    with open(path) as fh:
        for row in csv.reader(fh):
            if len(row) >= 4 and re.match(r"^\d\d/\d\d/\d{4}$", row[0]) and row[2]:
                out.append(dict(date=row[0], desc=row[1].strip(),
                                amt=float(row[2].replace(",", "")), ref=""))
    return out

def parse_pdf_text(path, year):
    """`pdftotext -layout eStmt.pdf out.txt` then feed out.txt here."""
    txt = open(path).read(); out = []
    dep = re.compile(r"^(\d\d/\d\d)/\d\d\s+(.+?)\s{2,}(?:(\d{6,})\s+)?([\d,]+\.\d\d)\s*$")
    for ln in txt.split("\n"):
        m = dep.match(ln.rstrip())
        if m and "Total" not in m.group(2):
            out.append(dict(date=f"{m.group(1)}/{year}", desc=re.sub(r"\s{2,}", " ", m.group(2)).strip(),
                            amt=float(m.group(4).replace(",", "")), ref=m.group(3) or ""))
    ck = re.compile(r"(\d\d/\d\d)\s+(\d{5,6})\*?\s+(\d{9,})\s+(-[\d,]+\.\d\d)")
    for ln in txt.split("\n"):
        for m in ck.finditer(ln):
            out.append(dict(date=f"{m.group(1)}/{year}", desc=f"Check {m.group(2)}",
                            amt=float(m.group(4).replace(",", "")), ref=m.group(3)))
    return out

def split_txns(txns):
    credits, numbered, unnumbered, other = [], [], [], []
    for t in txns:
        if t["amt"] > 0: credits.append(t); continue
        m = re.search(r"Check(?: Image)? (\d+)", t["desc"])
        if m: numbered.append((t, m.group(1)))
        elif t["desc"].strip() == "Check": unnumbered.append(t)
        else: other.append(t)
    return credits, numbered, unnumbered, other

# ---------------------------------------------------------------- matching
def match(rows, txns, month):
    res = dict(month=month, checks_ok=[], checks_ambiguous=[], checks_mismatch=[], checks_norow=[],
               unnumbered=[], other=[], dep_tie=[], dep_off=[], assign={})
    # idempotent re-runs: rows already tied to a DIFFERENT month are off the table;
    # rows already tied to THIS month count as tied and are reported, not re-matched.
    claimed  = {r["row"] for r in rows if r["stmtmo"] and str(r["stmtmo"]) != month}
    prior    = {r["row"] for r in rows if str(r["stmtmo"]) == month}
    claimed |= prior
    res["already"] = len(prior)
    credits, numbered, unnum, other = split_txns(txns)
    # non-check debits (transfers, sweeps, fees) must ALSO have a journal row.
    # Match on amount, and on the BoA confirmation number if it is in column D.
    res["other"], res["other_ok"] = [], []
    for t in other:
        conf = re.search(r"Confirmation#?\s*(\d+)", t["desc"])
        cands = [x for x in rows if x["dis"] and abs(x["dis"] - abs(t["amt"])) < 0.005
                 and (x["row"] in prior or x["row"] not in claimed)]
        if conf:
            byconf = [x for x in cands if conf.group(1) in str(x["ck"])]
            if byconf: cands = byconf
        if len(cands) == 1:
            res["other_ok"].append((t, cands[0]))
            if cands[0]["row"] not in prior:
                claimed.add(cands[0]["row"]); res["assign"][cands[0]["row"]] = (t["date"], month)
        else:
            res["other"].append((t, cands))
    res["totals"] = dict(credits=c2(sum(t["amt"] for t in credits)),
                         numbered=c2(sum(t["amt"] for t, _ in numbered)),
                         unnumbered=c2(sum(t["amt"] for t in unnum)),
                         other=c2(sum(t["amt"] for t in other)))

    by = defaultdict(list)
    for x in rows:
        if x["dis"] and x["ck"]: by[x["ck"]].append(x)
    for t, n in numbered:
        cands = [x for x in by.get(n, [])
                 if abs(x["dis"] - abs(t["amt"])) < 0.005 and x["row"] not in claimed]
        if len(cands) == 1:
            res["checks_ok"].append((t, n, cands[0]))
            claimed.add(cands[0]["row"]); res["assign"][cands[0]["row"]] = (t["date"], month)
        elif len(cands) > 1:
            # tie-break on the row's own date matching the clearing date
            same = [x for x in cands if fmt(x["date"]) == t["date"]]
            if len(same) == 1:
                res["checks_ok"].append((t, n, same[0]))
                claimed.add(same[0]["row"]); res["assign"][same[0]["row"]] = (t["date"], month)
            else:
                res["checks_ambiguous"].append((t, n, cands))
        elif any(x["row"] in prior and abs(x["dis"] - abs(t["amt"])) < 0.005 for x in by.get(n, [])):
            res["checks_ok"].append((t, n, next(x for x in by[n]
                                     if x["row"] in prior and abs(x["dis"] - abs(t["amt"])) < 0.005)))
        elif by.get(n):
            res["checks_mismatch"].append((t, n, by[n]))
        else:
            res["checks_norow"].append((t, n))

    # unnumbered checks: resolve by elimination on amount
    for t in unnum:
        done = [x for x in rows if x["row"] in prior and x["dis"]
                and abs(x["dis"] - abs(t["amt"])) < 0.005]
        if done:
            res["unnumbered"].append((t, done[:1])); continue
        cands = [x for x in rows if x["dis"] and x["row"] not in claimed
                 and abs(x["dis"] - abs(t["amt"])) < 0.005 and not _issued_after(x, t["date"])]
        res["unnumbered"].append((t, cands))
        if len(cands) == 1:
            claimed.add(cands[0]["row"]); res["assign"][cands[0]["row"]] = (t["date"], month)

    # deposits: group journal rows on their own date, compare each day's totals
    bank = defaultdict(float); cnt = defaultdict(int)
    for t in credits: bank[t["date"]] += t["amt"]; cnt[t["date"]] += 1
    jg = defaultdict(list)
    for x in rows:
        if x["dep"] and isinstance(x["date"], datetime.datetime) and (
                x["row"] not in claimed or x["row"] in prior):
            jg[fmt(x["date"])].append(x)
    for d in sorted(bank, key=lambda s: (s[6:], s[:2], s[3:5])):
        js, bs = c2(sum(x["dep"] for x in jg.get(d, []))), c2(bank[d])
        if abs(js - bs) < 0.005 and jg.get(d):
            res["dep_tie"].append((d, bs, cnt[d], jg[d]))
            for x in jg[d]:
                claimed.add(x["row"])
                if x["row"] not in prior: res["assign"][x["row"]] = (d, month)
        else:
            res["dep_off"].append((d, bs, cnt[d], js, jg.get(d, [])))
    return res

def _issued_after(x, datestr):
    """True if the journal row's check was written after `datestr` — it cannot have cleared then."""
    m, d, y = datestr.split("/")
    cleared = datetime.date(int(y), int(m), int(d))
    blob = f"{fmt(x['purpose'])} {fmt(x['notes'])}"
    mm = re.search(r"check (\d{1,2})/(\d{1,2})/(\d{4})", blob) or re.search(r"\((\d{1,2})/(\d{1,2})/(\d{4})\)", blob)
    if mm: issued = datetime.date(int(mm.group(3)), int(mm.group(1)), int(mm.group(2)))
    elif isinstance(x["date"], datetime.datetime): issued = x["date"].date()
    else: return False
    return issued > cleared

# ---------------------------------------------------------------- write / report
def write_assignments(assign, path=JOURNAL):
    ts = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
    bak = os.path.join(os.path.dirname(path), "Backups", f"Account-Journal_backup_{ts}_pre-recon.xlsx")
    os.makedirs(os.path.dirname(bak), exist_ok=True); shutil.copy2(path, bak)
    wb = openpyxl.load_workbook(path, data_only=False); ws = wb[SHEET]
    for r, (cd, mo) in assign.items():
        m, d, y = cd.split("/")
        cell = ws.cell(r, C["cleared"]); cell.value = datetime.datetime(int(y), int(m), int(d))
        cell.number_format = "M/D/YYYY"; cell.alignment = Alignment(horizontal="center")
        cell = ws.cell(r, C["stmtmo"]); cell.value = mo; cell.alignment = Alignment(horizontal="center")
    wb.save(path)
    return bak

def report(res, rows, bank_end=None):
    L = []; P = L.append
    t = res["totals"]
    P(f"# Statement {res['month']}")
    P(f"credits {t['credits']:,.2f} | numbered checks {t['numbered']:,.2f} | "
      f"unnumbered {t['unnumbered']:,.2f} | other {t['other']:,.2f}")
    P("")
    if res.get("already"): P(f"({res['already']} rows were already marked {res['month']} — re-run is idempotent)")
    P(f"## Checks: {len(res['checks_ok'])} tied, {len(res['checks_ambiguous'])} ambiguous, "
      f"{len(res['checks_mismatch'])} amount mismatch, {len(res['checks_norow'])} with NO journal row")
    for t_, n, cs in res["checks_ambiguous"]:
        P(f"  AMBIGUOUS ck {n} {t_['date']} {abs(t_['amt']):,.2f} -> rows "
          f"{[(c['row'], fmt(c['payee']), fmt(c['cli'])) for c in cs]}")
    for t_, n, cs in res["checks_mismatch"]:
        P(f"  AMOUNT MISMATCH ck {n} {t_['date']} bank {abs(t_['amt']):,.2f} vs journal "
          f"{[(c['row'], c['dis']) for c in cs]}")
    for t_, n in res["checks_norow"]:
        P(f"  *** NO JOURNAL ROW: ck {n} {t_['date']} {abs(t_['amt']):,.2f}")
    for t_, cs in res["unnumbered"]:
        if len(cs) == 1:
            P(f"  unnumbered {t_['date']} {abs(t_['amt']):,.2f} -> ck {cs[0]['ck']} r{cs[0]['row']} "
              f"({fmt(cs[0]['payee'])}) by elimination")
        else:
            P(f"  *** UNNUMBERED {t_['date']} {abs(t_['amt']):,.2f}: {len(cs)} candidates "
              f"{[(c['row'], c['ck'], fmt(c['payee'])) for c in cs]}")
    for t_, x in res.get("other_ok", []):
        P(f"  non-check debit {t_['date']} {abs(t_['amt']):,.2f} -> r{x['row']} "
          f"({fmt(x['purpose'])}) tied")
    for o, cands in res["other"]:
        P(f"  *** NON-CHECK DEBIT with NO journal row: {o['date']} {o['desc']} {o['amt']:,.2f}"
          + (f"  [{len(cands)} same-amount candidates: {[c['row'] for c in cands]}]" if cands else ""))
    P("")
    P("## Deposits")
    for d, bs, n, jr in res["dep_tie"]:
        P(f"  {d} {bs:>13,.2f} ({n} credits) = {len(jr)} journal rows  TIE")
    for d, bs, n, js, jr in res["dep_off"]:
        P(f"  *** {d} bank {bs:,.2f} ({n} credits) vs journal {js:,.2f} ({len(jr)} rows) "
          f"OFF by {bs - js:,.2f}")
        for x in jr: P(f"        r{x['row']} {x['dep']:,.2f} {fmt(x['cli'])}")
    P("")
    tD = sum(x["dep"] for x in rows); tS = sum(x["dis"] for x in rows)
    P(f"## Book\nBook balance (journal as it stands): {c2(tD - tS):,.2f}")
    if bank_end:
        P(f"Bank ending: {bank_end:,.2f}   gap: {c2(tD - tS - bank_end):,.2f}")
        P("Gap must decompose into: deposits in transit − outstanding checks + unrecorded bank debits.")
    return "\n".join(L)

# ---------------------------------------------------------------- main
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--stmt", required=True, help="statement CSV, or pdftotext -layout output")
    ap.add_argument("--month", required=True, help="e.g. 2026-09")
    ap.add_argument("--journal", default=JOURNAL)
    ap.add_argument("--bank-end", type=float, default=None)
    ap.add_argument("--write", action="store_true", help="backfill M/N (backs up first)")
    ap.add_argument("--json", default=None)
    a = ap.parse_args()

    if ensure_columns(a.journal): print("added columns M (Cleared Date) / N (Statement Month)")
    rows, last = load_journal(a.journal)
    print(f"journal: {len(rows)} rows (last data row {last})\n")
    txns = parse_pdf_text(a.stmt, a.month[:4]) if a.stmt.endswith(".txt") else parse_csv(a.stmt)
    res = match(rows, txns, a.month)
    print(report(res, rows, a.bank_end))
    if a.json:
        json.dump({str(k): v for k, v in res["assign"].items()}, open(a.json, "w"), indent=0)
    if a.write:
        bak = write_assignments(res["assign"], a.journal)
        print(f"\nWROTE {len(res['assign'])} rows. Backup: {bak}")
    else:
        print(f"\nDRY RUN — {len(res['assign'])} rows would be marked. Re-run with --write.")

if __name__ == "__main__":
    main()
