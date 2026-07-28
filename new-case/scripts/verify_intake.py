#!/usr/bin/env python3
"""
verify_intake.py — intake sheet 写后自检（闸二）

用途：填完 intake sheet 后立刻跑一遍，抓「值填错栏」这类静默错误。
原理：把每个值格和它左边的标签配对，按标签关键词推断该字段的**值类型**，
      再校验实际写入的值是否符合该类型。类型不符 → 报警。

历史事故（这个脚本就是为了防止它们再发生）：
  - I7  Policyholder ← 误填了保单期间 "03/22/2026-09/22/2026"
  - I8  Driver       ← 误填了 named insured（且括号里写着 "≠ driver"）
  - L12 Driver Phone ← 误填了地址（与 L13 重复）
  - F12 Child Seat   ← 误填了"是否预见撞击"的答案

用法：
    python3 verify_intake.py "<路径>/XXX Intake Sheet.xlsx" [--client-name "Qianxu Jin"]
"""

import re
import sys
from openpyxl import load_workbook

# (标签列, 值列, 区块名)
SECTIONS = [
    ("B", "C", "Clients Index"),
    ("E", "F", "Accident Info"),
    ("H", "I", "1P Insurance"),
    ("K", "L", "3P Insurance"),
    ("X", "Y", "Other Party Insurance"),
    ("AA", "AB", "Vehicle"),
]

# 允许出现在任何字段的"未填"占位
PLACEHOLDERS = {"", "pending", "n/a", "na", "none", "no", "yes", "tbd"}

DATE_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}")
PERIOD_RE = re.compile(r"\d{1,2}/\d{1,2}/\d{2,4}\s*-\s*\d{1,2}/\d{1,2}/\d{2,4}")
PHONE_RE = re.compile(r"(\d[\s\-().]*){10,}")
EMAIL_RE = re.compile(r"[^@\s]+@[^@\s]+\.[^@\s]+")
VIN_RE = re.compile(r"^[A-HJ-NPR-Z0-9]{17}$", re.I)
ZIP_RE = re.compile(r"\b\d{5}(-\d{4})?\b")
# 街道地址特征：数字开头 + 街道后缀
ADDR_RE = re.compile(
    r"\d+\s+\w+.*\b(st|street|ave|avenue|rd|road|dr|drive|blvd|ln|lane|ct|court|"
    r"way|pkwy|parkway|cir|circle|pl|place|ter|terrace|hwy)\b",
    re.I,
)


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def check(label, value, client_name=None):
    """返回 (级别, 说明) 列表；级别 = ERROR / WARN"""
    problems = []
    lab = norm(label).replace("\n", " ")
    val = str(value).strip() if value is not None else ""
    if norm(val) in PLACEHOLDERS:
        return problems

    # 字段名 = 标签第一行（\n 之后是填写提示，不能参与判定）
    lab_head = norm(label).split("\n")[0].strip()
    is_ = lambda *kw: any(k in lab_head for k in kw)
    # 这些字段虽含 driver/name 等词，但值不是姓名
    NOT_NAME = is_("dob", "birth", "license", "phone", "fax", "address", "email", "#")

    # —— 电话 / 传真 ——
    if is_("phone", "fax", "tel"):
        if ADDR_RE.search(val) or ZIP_RE.search(val):
            problems.append(("ERROR", "标签是电话/传真，值看起来是地址"))
        elif not PHONE_RE.search(val):
            problems.append(("WARN", "标签是电话/传真，值不像电话号码"))

    # —— 日期 / 期间 / 时间 ——
    if is_("period"):
        if not PERIOD_RE.search(val):
            problems.append(("WARN", "标签是 Period，值不像日期区间"))
    elif is_("time"):
        pass                                       # 时间格式照抄客人所填，不校验
    elif is_("date", "dob", "birth"):
        if not DATE_RE.search(val):
            problems.append(("WARN", "标签是日期，值不像日期"))

    # —— 姓名类字段不该是日期/地址/邮箱 ——
    if is_("policyholder", "insured", "owner", "name", "driver", "adjuster") and not NOT_NAME:
        if PERIOD_RE.search(val) or (DATE_RE.search(val) and len(val) < 30):
            problems.append(("ERROR", "标签是姓名类字段，值是日期/期间"))
        elif ADDR_RE.search(val):
            problems.append(("ERROR", "标签是姓名类字段，值看起来是地址"))
        elif EMAIL_RE.search(val):
            problems.append(("ERROR", "标签是姓名类字段，值是邮箱"))

    # —— 1P Driver 应当是我方客户 ——
    if client_name and lab_head.strip() == "driver":
        pass  # 区块级校验在 main 里做（只对 1P 生效）

    # —— 邮箱 ——
    if is_("email"):
        if not EMAIL_RE.search(val):
            problems.append(("WARN", "标签是 Email，值不含邮箱地址"))

    # —— VIN ——
    if is_("vin"):
        if not VIN_RE.match(val.replace(" ", "")):
            problems.append(("WARN", "VIN 不是 17 位"))

    # —— 地址字段反过来不该是电话 ——
    if is_("address"):
        if PHONE_RE.fullmatch(val.replace("-", "").replace(" ", "")):
            problems.append(("ERROR", "标签是地址，值看起来是电话号码"))

    # —— 夹带分析的信号词（照抄原则）——
    ANALYSIS = ["recommend", "confirm with", "apportionment", "⚠", "liability favors",
                "obtain ", "needs to", "should ", "per client", "note:"]
    if any(a in val.lower() for a in ANALYSIS) and len(val) > 60:
        problems.append(("WARN", "值疑似夹带分析/建议 —— intake 只记录客人陈述"))

    return problems


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    path = sys.argv[1]
    client_name = None
    if "--client-name" in sys.argv:
        client_name = sys.argv[sys.argv.index("--client-name") + 1]

    ws = load_workbook(path).active
    errors = warns = 0

    for lab_col, val_col, section in SECTIONS:
        printed = False
        for r in range(1, 75):
            label = ws[f"{lab_col}{r}"].value
            value = ws[f"{val_col}{r}"].value
            if not label or value in (None, ""):
                continue
            probs = check(label, value, client_name)

            # 1P Driver 必须是我方客户
            lab1 = norm(label).split("\n")[0].strip()
            if section == "1P Insurance" and lab1 == "driver" \
               and client_name and client_name.lower() not in str(value).lower():
                probs.append(("ERROR", f"1P Driver 应是实际开车的人（我方客户 {client_name}）"))

            for lvl, msg in probs:
                if not printed:
                    print(f"\n=== {section} ===")
                    printed = True
                tag = "❌ ERROR" if lvl == "ERROR" else "⚠️  WARN "
                lab_s = str(label).replace("\n", " / ")[:45]
                print(f"{tag} {val_col}{r} | {lab_s}\n         值: {str(value)[:90]}\n         → {msg}")
                if lvl == "ERROR":
                    errors += 1
                else:
                    warns += 1

    print(f"\n{'='*60}\n自检完成：{errors} 个 ERROR，{warns} 个 WARN")
    if errors:
        print("❌ 有 ERROR，必须修正后重跑。")
        sys.exit(2)
    print("✅ 无 ERROR。")


if __name__ == "__main__":
    main()
